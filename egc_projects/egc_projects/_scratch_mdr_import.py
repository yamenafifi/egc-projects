"""One-time import: PRJ2601049 (Siemens - PET CT - Abha)'s Master Document Register into real
EGC Project Document/Revision + EGC Submittal/Revision records, with the actual PDF attached and
a real response recorded through the same engine functions the Hub itself calls — not raw field
writes, so every invariant (immutable issued revisions, permanent submitted history, one writer
per derived field) holds exactly as it does for anything created through the UI.

Source: /tmp/project_import/Master_Document_Register.xlsx (sheet "MDR_Master") + PDFs in
/tmp/project_import/HI, matched to a row by filename prefix (the file name always starts with
the row's own "Doc Ref No").

Confirmed with the user before writing this:
- Approval Status codes: Code A -> Approved, Code B -> Approved with Comments,
  Code C -> Revise & Resubmit, Code D -> Rejected. "UR" (under review) and "Not Submitted" get
  no response recorded.
- Rows with no Doc Ref No ("-", 27 of them, all "Not Submitted" placeholders with no file) are
  skipped entirely.
- Rows with a real Doc Ref No but no matching PDF (7 of them) are skipped entirely.
- One EGC Submittal per document, one submission cycle, response recorded directly
  (submittal_control.record_response) - no invented multi-step reviewer workflow, since the MDR
  itself names no individual reviewers.

Idempotent: skips any Doc Ref No that already has an EGC Project Document on this project, so a
partial run (or a genuine re-run before applying to production) never double-creates.

    bench execute egc_projects.egc_projects._scratch_mdr_import.run
"""

from __future__ import annotations

import os
import re

import frappe
import openpyxl

from egc_projects.api import documents as documents_api
from egc_projects.api import submittals as submittals_api
from egc_projects.egc_projects import constants as c
from egc_projects.egc_projects import submittal_control

PROJECT = "PRJ2601049"
XLSX_PATH = "/tmp/project_import/Master_Document_Register.xlsx"
HI_DIR = "/tmp/project_import/HI"

DISCIPLINE_MAP = {
	"Architectural": "ARCH",
	"Civil": "CIVIL",
	"Electrical": "ELEC",
	"Mechanical": "MECH",
}

# MDR "Document Type" -> (EGC Document Type, EGC Submittal Type). Both master lists already
# exist and are seeded by install.py; nothing new added here.
DOC_TYPE_MAP = {
	"Material Submittal": ("Technical Data", "Material Submittal"),
	"Shop Drawing": ("Drawing", "Shop Drawing"),
	"Calculation": ("Calculation", "Calculation"),
	"Pre-Qulaification": ("Other", "Certificate"),
	"Proposed Drawing Submittal": ("Drawing", "Shop Drawing"),
}

STATUS_MAP = {
	"Code A": c.RESPONSE_APPROVED,
	"Code B": c.RESPONSE_APPROVED_WITH_COMMENTS,
	"Code C": c.RESPONSE_REVISE_AND_RESUBMIT,
	"Code D": c.RESPONSE_REJECTED,
	# "UR" (under review) and "Not Submitted" intentionally absent - no response recorded.
}

_HI_FILES = None


def _hi_files():
	global _HI_FILES
	if _HI_FILES is None:
		_HI_FILES = sorted(f for f in os.listdir(HI_DIR) if f.lower().endswith(".pdf"))
	return _HI_FILES


def _find_file(doc_ref: str) -> str | None:
	matches = [f for f in _hi_files() if f.startswith(doc_ref)]
	return matches[0] if matches else None


def _normalize_revision(value) -> str:
	# "Rev 00" -> "00", matching the app's own convention (see existing revisions in the DB).
	if not value:
		return "00"
	match = re.search(r"(\d+)", str(value))
	return match.group(1).zfill(2) if match else "00"


def _normalize_date(value):
	import datetime

	if value is None or value == "-":
		return None
	if isinstance(value, (datetime.date, datetime.datetime)):
		return value.date() if isinstance(value, datetime.datetime) else value
	# String dates in this sheet are DD/MM/YYYY.
	try:
		return datetime.datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
	except ValueError:
		return None


def _upload_file(doc_ref: str, filename: str) -> str:
	path = os.path.join(HI_DIR, filename)
	with open(path, "rb") as fh:
		content = fh.read()
	file_doc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": filename,
			"is_private": 1,
			"content": content,
		}
	)
	file_doc.insert(ignore_permissions=True)
	return file_doc.file_url


def _read_rows():
	wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
	ws = wb["MDR_Master"]
	rows = []
	for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
		if row[0] is None:
			continue
		rows.append(
			{
				"item_no": row[0],
				"doc_ref": row[1],
				"title": row[2],
				"discipline": row[3],
				"doc_type": row[4],
				"revision": row[5],
				"submission_date": row[6],
				"actual_reply_date": row[9],
				"status": row[10],
				"remarks": row[16],
			}
		)
	return rows


def run() -> None:
	frappe.set_user("Administrator")
	rows = _read_rows()

	skipped_no_ref = []
	skipped_no_file = []
	imported = []
	already_present = []
	failed = []

	for row in rows:
		doc_ref = row["doc_ref"]
		if not doc_ref or doc_ref == "-":
			skipped_no_ref.append(row)
			continue

		filename = _find_file(doc_ref)
		if not filename:
			skipped_no_file.append(row)
			continue

		if frappe.db.exists("EGC Project Document", {"project": PROJECT, "document_number": doc_ref}):
			already_present.append(doc_ref)
			continue

		try:
			_import_row(row, doc_ref, filename)
			imported.append(doc_ref)
			frappe.db.commit()
		except Exception as e:
			frappe.db.rollback()
			failed.append((doc_ref, str(e)))

	print("\n=== MDR Import Report ===")
	print(f"Imported: {len(imported)}")
	for ref in imported:
		print(f"  + {ref}")
	print(f"Already present (skipped): {len(already_present)}")
	for ref in already_present:
		print(f"  = {ref}")
	print(f"Failed: {len(failed)}")
	for ref, err in failed:
		print(f"  ! {ref}: {err}")
	print(f"Skipped - no Doc Ref No ({len(skipped_no_ref)} placeholder rows, not imported):")
	for row in skipped_no_ref:
		print(f"  - {row['title']}")
	print(f"Skipped - no matching PDF ({len(skipped_no_file)} rows, not imported):")
	for row in skipped_no_file:
		print(f"  - {row['doc_ref']}: {row['title']}")


def _import_row(row: dict, doc_ref: str, filename: str) -> None:
	discipline = DISCIPLINE_MAP.get(row["discipline"])
	doc_type, submittal_type = DOC_TYPE_MAP.get(row["doc_type"], ("Other", "Certificate"))
	revision_label = _normalize_revision(row["revision"])
	submission_date = _normalize_date(row["submission_date"])
	reply_date = _normalize_date(row["actual_reply_date"])
	title = (row["title"] or doc_ref).strip()

	file_url = _upload_file(doc_ref, filename)

	document = documents_api.create_document(
		project=PROJECT,
		document_number=doc_ref,
		title=title,
		document_type=doc_type,
		discipline=discipline,
	)
	revision = documents_api.create_document_revision(
		document=document["name"],
		revision=revision_label,
		file=file_url,
		revision_date=submission_date,
		remarks=row["remarks"],
	)
	documents_api.submit_document_revision(revision["name"])

	submittal = submittals_api.create_submittal(
		project=PROJECT,
		submittal_number=doc_ref,
		title=title,
		submittal_type=submittal_type,
		discipline=discipline,
	)
	submission = submittals_api.create_first_submission(submittal["name"], revision_label=revision_label)
	submittals_api.add_submission_document(submission["name"], revision["name"])
	submittals_api.submit_submission(submission["name"])

	status = (row["status"] or "").strip()
	response = STATUS_MAP.get(status)
	if response:
		submittal_control.record_response(
			submission["name"],
			response,
			remarks=row["remarks"],
			response_date=reply_date or submission_date,
		)
